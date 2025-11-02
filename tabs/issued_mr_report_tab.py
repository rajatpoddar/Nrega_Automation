# tabs/issued_mr_report_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, re, json
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

from fpdf import FPDF
from PIL import Image, ImageDraw, ImageFont # <-- Added PIL imports
from utils import resource_path
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry
import config

class IssuedMrReportTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        super().__init__(parent, app_instance, automation_key="issued_mr_report")
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) # Main notebook takes up space
        
        # Headers based on report details page.htm
        self.report_headers = [
            "S No.", "Panchayat", "Work Code", "Work Name", 
            "Work Category", "Work Type", "Agency Name"
        ]
        
        self._create_widgets()
        self.load_inputs()

    def _create_widgets(self):
        # Frame for all user input controls
        controls_frame = ctk.CTkFrame(self)
        controls_frame.grid(row=0, column=0, sticky="new", padx=10, pady=10)
        controls_frame.grid_columnconfigure(1, weight=1)

        # --- Input Fields ---
        ctk.CTkLabel(controls_frame, text="State:").grid(row=0, column=0, sticky='w', padx=15, pady=(15, 5))
        self.state_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("issued_mr_state"),
                                             app_instance=self.app, history_key="issued_mr_state")
        self.state_entry.grid(row=0, column=1, sticky='ew', padx=15, pady=(15, 5))

        ctk.CTkLabel(controls_frame, text="District:").grid(row=1, column=0, sticky='w', padx=15, pady=5)
        self.district_entry = AutocompleteEntry(controls_frame, 
                                                suggestions_list=self.app.history_manager.get_suggestions("issued_mr_district"),
                                                app_instance=self.app, history_key="issued_mr_district")
        self.district_entry.grid(row=1, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Block:").grid(row=2, column=0, sticky='w', padx=15, pady=5)
        self.block_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("issued_mr_block"),
                                             app_instance=self.app, history_key="issued_mr_block")
        self.block_entry.grid(row=2, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Panchayat:").grid(row=3, column=0, sticky='w', padx=15, pady=5)
        self.panchayat_entry = AutocompleteEntry(controls_frame, 
                                                 suggestions_list=self.app.history_manager.get_suggestions("issued_mr_panchayat"),
                                                 app_instance=self.app, history_key="issued_mr_panchayat")
        self.panchayat_entry.grid(row=3, column=1, sticky='ew', padx=15, pady=5)

        # Removed Delay Column

        action_frame = self._create_action_buttons(parent_frame=controls_frame)
        action_frame.grid(row=4, column=0, columnspan=2, pady=10) # Row updated to 4

        # --- Output Tabs ---
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        workcode_tab = notebook.add("Workcode List")
        results_tab = notebook.add("Results Table")
        self._create_log_and_status_area(parent_notebook=notebook)

        # 1. Workcode List Tab
        workcode_tab.grid_columnconfigure(0, weight=1)
        workcode_tab.grid_rowconfigure(1, weight=1)
        
        copy_frame = ctk.CTkFrame(workcode_tab, fg_color="transparent")
        copy_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        self.copy_wc_button = ctk.CTkButton(copy_frame, text="Copy Workcodes", command=self._copy_workcodes)
        self.copy_wc_button.pack(side="left")

        # --- MODIFIED BUTTON ---
        self.run_dup_mr_button = ctk.CTkButton(copy_frame,
                                                  text="Run Duplicate MR Print",
                                                  command=self._run_duplicate_mr,
                                                  fg_color="#D35400", # Orange
                                                  hover_color="#E67E22")
        self.run_dup_mr_button.pack_forget() # Hide it initially
        # --- END MODIFIED BUTTON ---

        self.workcode_textbox = ctk.CTkTextbox(workcode_tab, state="disabled")
        self.workcode_textbox.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        # 2. Results Tab (Table)
        results_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)
        
        export_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        export_frame.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.export_button = ctk.CTkButton(export_frame, text="Export Report", command=self.export_report)
        self.export_button.pack(side="left")
        
        self.export_format_menu = ctk.CTkOptionMenu(export_frame, values=["Excel (.xlsx)", "PDF (.pdf)", "PNG (.png)"])
        self.export_format_menu.pack(side="left", padx=5)

        # --- Update Treeview columns and widths ---
        self.results_tree = ttk.Treeview(results_tab, columns=self.report_headers, show='headings')
        for col in self.report_headers: 
            self.results_tree.heading(col, text=col)
            
        # Adjust column widths based on new headers
        self.results_tree.column("S No.", width=40, anchor='center')
        self.results_tree.column("Panchayat", width=100)
        self.results_tree.column("Work Code", width=200)
        self.results_tree.column("Work Name", width=350)
        self.results_tree.column("Work Category", width=150)
        self.results_tree.column("Work Type", width=150)
        self.results_tree.column("Agency Name", width=100)
        # --- END ---

        self.results_tree.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree)

    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        self.state_entry.configure(state=state)
        self.district_entry.configure(state=state)
        self.block_entry.configure(state=state)
        self.panchayat_entry.configure(state=state)
        self.run_dup_mr_button.configure(state=state)

    def reset_ui(self):
        # Clear inputs
        self.state_entry.delete(0, tkinter.END)
        self.district_entry.delete(0, tkinter.END)
        self.block_entry.delete(0, tkinter.END)
        self.panchayat_entry.delete(0, tkinter.END)
        
        # Clear results
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        self._update_workcode_textbox("")
        
        self.app.log_message(self.log_display, "Form has been reset.")
        self.update_status("Ready", 0.0)
        
    def start_automation(self):
        self.run_dup_mr_button.pack_forget()
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        self._update_workcode_textbox("") # Clear workcode list
        
        inputs = {
            'state': self.state_entry.get().strip(), 
            'district': self.district_entry.get().strip(), 
            'block': self.block_entry.get().strip(),
            'panchayat': self.panchayat_entry.get().strip(),
        }
        
        if not all([inputs['state'], inputs['district'], inputs['block'], inputs['panchayat']]):
            messagebox.showwarning("Input Error", "All fields are required."); return
        
        self.save_inputs(inputs)
        self.app.update_history("issued_mr_state", inputs['state'])
        self.app.update_history("issued_mr_district", inputs['district'])
        self.app.update_history("issued_mr_block", inputs['block'])
        self.app.update_history("issued_mr_panchayat", inputs['panchayat'])
        
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(inputs,))

    def _solve_captcha(self, driver, wait):
        """Solves the math CAPTCHA on the MIS report page."""
        self.app.log_message(self.log_display, "Attempting to solve CAPTCHA...")
        captcha_label_id = "ContentPlaceHolder1_lblStopSpam"; captcha_textbox_id = "ContentPlaceHolder1_txtCaptcha"; verify_button_id = "ContentPlaceHolder1_btnLogin"
        try:
            captcha_element = wait.until(EC.presence_of_element_located((By.ID, captcha_label_id)))
            captcha_text = captcha_element.text
            match = re.search(r'(\d+)\s*([+\-*])\s*(\d+)', captcha_text)
            if not match: raise ValueError("Could not parse CAPTCHA expression.")
            num1, operator, num2 = match.groups(); num1, num2 = int(num1), int(num2)
            result = 0
            if operator == '+': result = num1 + num2
            elif operator == '-': result = num1 - num2
            elif operator == '*': result = num1 * num2
            self.app.log_message(self.log_display, f"Solved: {captcha_text.strip()} = {result}")
            driver.find_element(By.ID, captcha_textbox_id).send_keys(str(result))
            driver.find_element(By.ID, verify_button_id).click()
            # Wait briefly to see if an error message appears
            time.sleep(1)
            if "Invalid Captcha Code" in driver.page_source:
                raise ValueError("CAPTCHA verification failed.")
            return True
        except TimeoutException:
            # If CAPTCHA elements are not found, assume it's already solved or not present
            self.app.log_message(self.log_display, "CAPTCHA not found or already bypassed.", "info")
            return True # Continue as if successful
        except ValueError as e:
            self.app.log_message(self.log_display, f"CAPTCHA Error: {e}", "error")
            raise # Re-raise the error to be caught by the main logic

    def run_automation_logic(self, inputs, retries=1):
        self.app.after(0, self.set_ui_state, True)
        self.app.after(0, self.app.set_status, "Starting Issued MR Report...") 
        self.app.after(0, self.update_status, "Initializing...", 0.0)
        self.app.clear_log(self.log_display)
        self.app.log_message(self.log_display, "Starting Issued MR Report automation...")

        try:
            driver = self.app.get_driver()
            if not driver:
                self.app.after(0, self.app.set_status, "Browser not found")
                return 

            wait = WebDriverWait(driver, 20)

            self.app.after(0, self.app.set_status, "Navigating to MIS portal...")
            self.app.after(0, self.update_status, "Navigating...", 0.05)
            self.app.log_message(self.log_display, "Navigating to MIS portal...")
            driver.get(config.MIS_REPORTS_CONFIG["base_url"])

            self.app.after(0, self.app.set_status, "Solving CAPTCHA...")
            self.app.after(0, self.update_status, "Solving CAPTCHA...", 0.1)
            self._solve_captcha(driver, wait) # Handles potential failure
            self.app.log_message(self.log_display, "CAPTCHA step passed. Selecting state...")

            self.app.after(0, self.app.set_status, f"Selecting State: {inputs['state']}...")
            self.app.after(0, self.update_status, "Selecting State...", 0.15)
            state_select = wait.until(EC.element_to_be_clickable((By.ID, "ContentPlaceHolder1_ddl_States")))
            Select(state_select).select_by_visible_text(inputs['state'].upper())
            wait.until(EC.presence_of_element_located((By.LINK_TEXT, "Dashboard for Delay Monitoring System")))

            self.app.after(0, self.app.set_status, "Opening Issued MR Report...")
            self.app.after(0, self.update_status, "Opening Report...", 0.2)
            self.app.log_message(self.log_display, "Clicking 'MGNREGS daily status as per e-muster issued'...")
            report_link_text = "MGNREGS daily status as per e-muster issued"
            report_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, report_link_text)))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", report_link)
            time.sleep(1); report_link.click()

            self.app.after(0, self.app.set_status, f"Selecting District: {inputs['district']}...")
            self.app.after(0, self.update_status, "Selecting District...", 0.25)
            self.app.log_message(self.log_display, f"Drilling down to District: {inputs['district']}")
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['district'].upper()))).click()

            self.app.after(0, self.app.set_status, f"Selecting Block: {inputs['block']}...")
            self.app.after(0, self.update_status, "Selecting Block...", 0.3)
            self.app.log_message(self.log_display, f"Drilling down to Block: {inputs['block']}")
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['block'].upper()))).click()

            # --- MODIFIED LOGIC: Click on "No. Of Ongoing Works..." column ---
            self.app.after(0, self.app.set_status, f"Finding Panchayat: {inputs['panchayat']}...")
            self.app.after(0, self.update_status, "Finding Panchayat...", 0.35)
            self.app.log_message(self.log_display, f"Finding Panchayat row: {inputs['panchayat']}")
            
            # This XPath targets the main table from 'issue mr report.htm'
            main_table_xpath = "//table[.//b[text()='SNo.'] and .//b[text()='Panchayats']]"
            wait.until(EC.presence_of_element_located((By.XPATH, f"{main_table_xpath}//tr[1]/td/b[text()='Panchayats']")))

            panchayat_row_xpath = f"{main_table_xpath}//tr[td[2][normalize-space()='{inputs['panchayat']}']]"
            panchayat_row = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, panchayat_row_xpath)))

            self.app.after(0, self.app.set_status, "Clicking Report Link...")
            self.app.after(0, self.update_status, "Clicking Link...", 0.45)
            self.app.log_message(self.log_display, "Clicking 'No. Of Ongoing Works on which MR Issued' (Column 6)...")
            
            # Column 6 (index 5) holds the link we need
            target_cell = panchayat_row.find_element(By.XPATH, "./td[6]")

            try:
                target_link = target_cell.find_element(By.TAG_NAME, "a")
                link_text = target_link.text.strip()
                if link_text == '0':
                    self.app.log_message(self.log_display, "Column has value 0. No data to fetch.", "warning")
                    messagebox.showinfo("No Data", f"The 'Ongoing Works' column has a value of 0 for {inputs['panchayat']}. No details to display.")
                    self.success_message = None
                    self.app.after(0, self.app.set_status, "No data found")
                    return

                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_link)
                time.sleep(0.5)
                target_link.click()

            except NoSuchElementException:
                 cell_text = target_cell.text.strip()
                 if cell_text == '0':
                    self.app.log_message(self.log_display, "Column has value 0 (not a link). No data to fetch.", "warning")
                    messagebox.showinfo("No Data", f"The 'Ongoing Works' column has a value of 0 for {inputs['panchayat']}. No details to display.")
                    self.success_message = None
                    self.app.after(0, self.app.set_status, "No data found")
                    return
                 else:
                    raise ValueError(f"Target cell for 'Ongoing Works' does not contain a clickable link (text: {cell_text}).")

            # --- END MODIFIED LOGIC ---

            self.app.after(0, self.app.set_status, "Loading Final Report...")
            self.app.after(0, self.update_status, "Loading Final Report...", 0.5)
            self.app.log_message(self.log_display, "Waiting for final report table...")
            
            # This XPath targets the final table from 'report details page.htm'
            FINAL_TABLE_XPATH = "//table[@align='center' and .//b[text()='Work Code']]"
            table = wait.until(EC.presence_of_element_located((By.XPATH, FINAL_TABLE_XPATH)))
            rows = table.find_elements(By.XPATH, ".//tr[position()>1]") # Skip header row

            total_rows = len(rows)
            if total_rows == 0:
                self.app.log_message(self.log_display, "Final report table is empty.", "warning")
                messagebox.showinfo("No Data", f"No detailed records found for {inputs['panchayat']}.")
                self.success_message = None
                self.app.after(0, self.app.set_status, "No data found")
                return

            self.app.log_message(self.log_display, f"Found {total_rows} records in the final table. Processing...")

            workcode_list = []
            scraped_mr_count = 0

            for i, row in enumerate(rows):
                if self.app.stop_events[self.automation_key].is_set():
                    self.app.log_message(self.log_display, "Stop signal received.", "warning")
                    break 

                progress = 0.5 + ( (i + 1) / total_rows ) * 0.45
                status_msg = f"Processing row {i+1}/{total_rows}"
                self.app.after(0, self.app.set_status, status_msg)
                self.app.after(0, self.update_status, status_msg, progress)

                cells = row.find_elements(By.TAG_NAME, "td")
                if not cells or len(cells) < len(self.report_headers):
                    self.app.log_message(self.log_display, f"Skipping row {i+1}, expected at least {len(self.report_headers)} columns, found {len(cells)}.", "warning")
                    continue

                scraped_data = [cell.text.strip() for cell in cells[:len(self.report_headers)]]
                work_code = scraped_data[2] # Column 3 is "Work Code"
                scraped_mr_count += 1
                row_data = tuple(scraped_data)

                self.app.after(0, lambda data=row_data: self.results_tree.insert("", "end", values=data))
                if work_code:
                    workcode_list.append(work_code)

            if self.app.stop_events[self.automation_key].is_set():
                 self.app.log_message(self.log_display, "Automation stopped by user.", "warning")
                 self.success_message = None 
                 return 

            # Update workcode list
            unique_workcodes = list(dict.fromkeys(workcode_list))
            self.app.after(0, self._update_workcode_textbox, "\n".join(unique_workcodes))

            self.app.log_message(self.log_display, f"Processing complete. Found {scraped_mr_count} MRs listed.", "success")
            self.success_message = f"Issued MR Report automation has finished.\nFound {scraped_mr_count} Issued MRs in {inputs['panchayat']}."

        except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
            if "Session Expired" in driver.page_source and retries > 0:
                self.app.log_message(self.log_display, "Session expired, attempting retry...", "warning")
                self.app.after(0, self.app.set_status, "Session expired, retrying...")
                self.app.after(0, self.update_status, "Retrying...", 0.0)
                self.run_automation_logic(inputs, retries - 1)
                return 
            error_msg = f"A browser error occurred: {str(e).splitlines()[0]}"
            self.app.log_message(self.log_display, error_msg, "error")
            messagebox.showerror("Automation Error", error_msg)
            self.app.after(0, self.app.set_status, "Browser Error")
            self.success_message = None
        except ValueError as e: 
             error_msg = f"Data processing error: {e}"
             self.app.log_message(self.log_display, error_msg, "error")
             messagebox.showerror("Automation Error", error_msg)
             self.app.after(0, self.app.set_status, "Data Error")
             self.success_message = None
        except Exception as e:
            self.app.log_message(self.log_display, f"An unexpected error occurred: {e}", "error")
            messagebox.showerror("Critical Error", f"An unexpected error occurred: {e}")
            self.app.after(0, self.app.set_status, "Unexpected Error")
            self.success_message = None
        finally:
            self.app.after(0, self.set_ui_state, False)
            
            final_app_status = "Automation Stopped" if self.app.stop_events[self.automation_key].is_set() else \
                              ("Automation Finished" if hasattr(self, 'success_message') and self.success_message else "Automation Failed")
            final_tab_status = "Stopped" if self.app.stop_events[self.automation_key].is_set() else \
                              ("Finished" if hasattr(self, 'success_message') and self.success_message else "Failed")

            self.app.after(0, self.app.set_status, final_app_status)
            self.app.after(0, self.update_status, final_tab_status, 1.0)

            if not self.app.stop_events[self.automation_key].is_set():
                 self.app.after(5000, lambda: self.app.set_status("Ready")) 
                 self.app.after(5000, lambda: self.update_status("Ready", 0.0))

            if hasattr(self, 'success_message') and self.success_message and not self.app.stop_events[self.automation_key].is_set():
                self.app.after(100, lambda: messagebox.showinfo("Complete", self.success_message))
                # Show the button to transfer data
                self.app.after(0, lambda: self.run_dup_mr_button.pack(side="left", padx=(10, 0)))

    def _update_workcode_textbox(self, text):
        self.workcode_textbox.configure(state="normal")
        self.workcode_textbox.delete("1.0", tkinter.END)
        self.workcode_textbox.insert("1.0", text)
        self.workcode_textbox.configure(state="disabled")

    def _copy_workcodes(self):
        text = self.workcode_textbox.get("1.0", tkinter.END).strip()
        if text:
            self.app.clipboard_clear()
            self.app.clipboard_append(text)
            messagebox.showinfo("Copied", f"{len(text.splitlines())} workcodes copied to clipboard.", parent=self)
        else:
            messagebox.showwarning("Empty", "There are no workcodes to copy.", parent=self)

    # --- Renamed method ---
    def _run_duplicate_mr(self):
        """Called when the 'Run Duplicate MR Print' button is clicked."""
        workcodes = self.workcode_textbox.get("1.0", tkinter.END).strip()
        panchayat_name = self.panchayat_entry.get().strip()

        if not workcodes:
            messagebox.showwarning("No Data", "There are no workcodes to send to the Duplicate MR Print tab.", parent=self)
            return
        
        if not panchayat_name:
            messagebox.showwarning("No Data", "Panchayat name is missing. Cannot send to Duplicate MR Print tab.", parent=self)
            return

        # Call the new method in the main app
        self.app.switch_to_duplicate_mr_with_data(workcodes, panchayat_name)
    # --- END Renamed method ---

    def export_report(self):
        if not self.results_tree.get_children():
            messagebox.showinfo("No Data", "There are no results to export.")
            return

        panchayat = self.panchayat_entry.get().strip() or "Report"
        safe_panchayat = re.sub(r'[\\/*?:"<>|]', '_', panchayat) 
        
        export_format = self.export_format_menu.get()
        current_year = datetime.now().strftime("%Y")
        current_date_str = datetime.now().strftime("%d-%b-%Y") 

        headers = self.report_headers
        data = [self.results_tree.item(item, 'values') for item in self.results_tree.get_children()]

        pdf_title = f"Issued MR Report - {panchayat}" 
        date_str_header = f"Date - {datetime.now().strftime('%d-%m-%Y')}" 
        excel_title = f"{pdf_title} {date_str_header}" 
        
        downloads_path = self.app.get_user_downloads_path()
        target_dir = os.path.join(downloads_path, f"Reports {current_year}", safe_panchayat)
        try:
            os.makedirs(target_dir, exist_ok=True)
        except OSError as e:
             messagebox.showerror("Folder Error", f"Could not create report directory:\n{target_dir}\nError: {e}")
             return

        if "Excel" in export_format:
            ext = ".xlsx"
            file_type_tuple = ("Excel Workbook", "*.xlsx")
            default_filename = f"Issued_MR_Report_{safe_panchayat}-{current_date_str}{ext}"
        elif "PDF" in export_format:
            ext = ".pdf"
            file_type_tuple = ("PDF Document", "*.pdf")
            default_filename = f"Issued_MR_Report_{safe_panchayat}-{current_date_str}{ext}"
        elif "PNG" in export_format: 
            ext = ".png"
            file_type_tuple = ("PNG Image", "*.png")
            default_filename = f"Issued_MR_Report_{safe_panchayat}-{current_date_str}{ext}"
        else:
            return 

        file_path = filedialog.asksaveasfilename(
            initialdir=target_dir, 
            initialfile=default_filename, 
            defaultextension=ext,
            filetypes=[file_type_tuple, ("All Files", "*.*")],
            title="Save Report As"
        )
        
        if not file_path: return 

        if "Excel" in export_format:
            success = self._save_to_excel(data, headers, excel_title, file_path) 
            if success:
                messagebox.showinfo("Success", f"Excel report saved successfully to:\n{file_path}")
        
        elif "PDF" in export_format:
            # SNo, Panchayat, Work Code, Work Name, Category, Type, Agency
            col_widths = [12, 30, 60, 100, 40, 40, 30] 
            total_width_ratio = sum(col_widths)
            effective_page_width = 297 - 20 
            actual_col_widths = [(w / total_width_ratio) * effective_page_width for w in col_widths]
            
            success = self.generate_report_pdf(data, headers, actual_col_widths, pdf_title, date_str_header, file_path) 
            if success:
                messagebox.showinfo("Success", f"PDF report saved successfully to:\n{file_path}")

        elif "PNG" in export_format: 
            success = self._save_to_png(data, headers, pdf_title, date_str_header, file_path)
            if success:
                messagebox.showinfo("Success", f"PNG report saved successfully to:\n{file_path}")


    def _save_to_excel(self, data, headers, title, file_path):
        try:
            df = pd.DataFrame(data, columns=headers)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                sheet_name = 'Issued MR Report'
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                worksheet = writer.sheets[sheet_name]
                
                worksheet['A1'] = title
                worksheet['A1'].font = Font(bold=True, size=14)
                worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers)) 
                
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
                for cell in worksheet["2:2"]:
                    cell.font = header_font
                    cell.fill = header_fill

                for col_idx, col in enumerate(df.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    try:
                        max_length = max(len(str(col)), df[col].astype(str).map(len).max())
                    except (TypeError, ValueError):
                         max_length = len(str(col)) 
                    adjusted_width = min((max_length + 2), 50) 
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            return True
        except Exception as e:
            messagebox.showerror("Excel Export Error", f"Could not generate Excel report.\nError: {e}", parent=self)
            return False

    # This is the PDF generation from base_tab, included here for completeness
    def generate_report_pdf(self, data, headers, col_widths, title, date_str, file_path):
        return super().generate_report_pdf(data, headers, col_widths, title, date_str, file_path)

    # This is the PNG generation from base_tab, included here for completeness
    def _save_to_png(self, data, headers, title, date_str, file_path):
        # SNo, Panchayat, Work Code, Work Name, Category, Type, Agency
        base_col_widths = [0.05, 0.10, 0.20, 0.30, 0.15, 0.15, 0.05]
        
        try:
            font_path_regular = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
            font_path_bold = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
            font_title = ImageFont.truetype(font_path_bold, 28)
            font_date = ImageFont.truetype(font_path_regular, 18)
            font_header = ImageFont.truetype(font_path_bold, 16)
            font_body = ImageFont.truetype(font_path_regular, 14)
        except IOError:
            font_title = ImageFont.load_default(size=28)
            font_date = ImageFont.load_default(size=18)
            font_header = ImageFont.load_default(size=16)
            font_body = ImageFont.load_default(size=14)
        
        img_width = 2400; margin_x = 80; margin_y = 60
        header_bg_color = (220, 235, 255); row_even_bg_color = (255, 255, 255); row_odd_bg_color = (245, 245, 245); text_color = (0, 0, 0); border_color = (180, 180, 180)

        available_width = img_width - (2 * margin_x)
        col_widths_pixels = [w * available_width for w in base_col_widths]

        for i, header in enumerate(headers):
            header_width = font_header.getlength(header) + 40
            if col_widths_pixels[i] < header_width:
                diff = header_width - col_widths_pixels[i]
                col_widths_pixels[i] = header_width
                col_widths_pixels[3] -= diff # Steal from Work Name (index 3)

        if col_widths_pixels[3] < 100: col_widths_pixels[3] = 100
        
        current_total_width = sum(col_widths_pixels)
        scale_factor = available_width / current_total_width
        col_widths_pixels = [w * scale_factor for w in col_widths_pixels]

        initial_height = 1600
        img = Image.new("RGB", (img_width, initial_height), (255, 255, 255))
        draw = ImageDraw.Draw(img)
        current_y = margin_y
        
        title_bbox = font_title.getbbox(title); title_height = title_bbox[3] - title_bbox[1]
        title_text_width = font_title.getlength(title); title_x = (img_width - title_text_width) / 2
        draw.text((title_x, current_y), title, font=font_title, fill=text_color)
        current_y += title_height + 5

        date_bbox = font_date.getbbox(date_str); date_height = date_bbox[3] - date_bbox[1]
        date_text_width = font_date.getlength(date_str); date_x = img_width - margin_x - date_text_width
        draw.text((date_x, current_y), date_str, font=font_date, fill=text_color)
        current_y += date_height + 20

        header_y_start = current_y; header_height = 0
        for i, header in enumerate(headers):
            wrapped_header = self._wrap_text(header, font_header, col_widths_pixels[i] - 10)
            line_height = (font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1]) * 1.2
            header_height = max(header_height, len(wrapped_header) * line_height + 10)
        
        current_x = margin_x
        for i, header in enumerate(headers):
            draw.rectangle([current_x, header_y_start, current_x + col_widths_pixels[i], header_y_start + header_height], fill=header_bg_color, outline=border_color, width=1)
            wrapped_header = self._wrap_text(header, font_header, col_widths_pixels[i] - 20)
            line_height = (font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1]) * 1.2
            total_text_height = len(wrapped_header) * line_height
            text_y = header_y_start + (header_height - total_text_height) / 2
            
            for line in wrapped_header:
                line_width = font_header.getlength(line)
                draw.text((current_x + (col_widths_pixels[i] - line_width) / 2, text_y), line, font=font_header, fill=text_color)
                text_y += line_height
            current_x += col_widths_pixels[i]
        current_y += header_height

        line_height = (font_body.getbbox("Tg")[3] - font_body.getbbox("Tg")[1]) * 1.2
        for row_idx, row_data in enumerate(data):
            row_bg_color = row_even_bg_color if row_idx % 2 == 0 else row_odd_bg_color
            max_row_text_height = 0
            temp_wrapped_cells = []
            for i, cell_text in enumerate(row_data):
                wrapped_lines = self._wrap_text(str(cell_text), font_body, col_widths_pixels[i] - 20)
                temp_wrapped_cells.append(wrapped_lines)
                max_row_text_height = max(max_row_text_height, len(wrapped_lines) * line_height)

            row_data_height = max_row_text_height + 10
            if current_y + row_data_height + margin_y > img.height:
                new_height = int(img.height + (row_data_height + margin_y) * 10)
                new_img = Image.new("RGB", (img_width, new_height), (255, 255, 255))
                new_img.paste(img, (0, 0))
                img = new_img
                draw = ImageDraw.Draw(img)

            current_x = margin_x
            for i, cell_text in enumerate(row_data):
                draw.rectangle([current_x, current_y, current_x + col_widths_pixels[i], current_y + row_data_height], fill=row_bg_color, outline=border_color, width=1)
                wrapped_lines = temp_wrapped_cells[i]
                text_y = current_y + 5
                for line in wrapped_lines:
                    draw.text((current_x + 10, text_y), line, font=font_body, fill=text_color)
                    text_y += line_height
                current_x += col_widths_pixels[i]
            current_y += row_data_height

        current_y += 15
        footer_text = "Report Generated by NregaBot.com"; footer_font = font_body
        footer_bbox = footer_font.getbbox(footer_text); footer_height = footer_bbox[3] - footer_bbox[1]
        footer_y_pos = current_y + 10

        if footer_y_pos + footer_height + margin_y > img.height:
            new_height = int(footer_y_pos + footer_height + margin_y)
            new_img = Image.new("RGB", (img_width, new_height), (255, 255, 255))
            new_img.paste(img, (0, 0))
            img = new_img
            draw = ImageDraw.Draw(img)
        
        draw.text((margin_x, footer_y_pos), footer_text, font=footer_font, fill=text_color)
        current_y = footer_y_pos + footer_height

        final_img = img.crop((0, 0, img_width, current_y + margin_y))
        final_img.save(file_path, "PNG", dpi=(300, 300))
        return True

    # This is the wrap_text from base_tab, included here for completeness
    def _wrap_text(self, text, font, max_width):
        return super()._wrap_text(text, font, max_width)

        
    def save_inputs(self, inputs):
        save_data = {k: inputs.get(k) for k in ('state', 'district', 'block', 'panchayat')}
        try:
            config_file = self.app.get_data_path("issued_mr_report_inputs.json")
            with open(config_file, 'w') as f:
                json.dump(save_data, f, indent=4)
        except Exception as e:
            print(f"Error saving Issued MR Report inputs: {e}")

    def load_inputs(self):
        try:
            config_file = self.app.get_data_path("issued_mr_report_inputs.json")
            if not os.path.exists(config_file): return
            
            with open(config_file, 'r') as f: data = json.load(f)
            
            self.state_entry.delete(0, 'end')
            self.state_entry.insert(0, data.get('state', ''))
            self.district_entry.delete(0, 'end')
            self.district_entry.insert(0, data.get('district', ''))
            self.block_entry.delete(0, 'end')
            self.block_entry.insert(0, data.get('block', ''))
            self.panchayat_entry.delete(0, 'end')
            self.panchayat_entry.insert(0, data.get('panchayat', ''))
        except Exception as e:
            print(f"Error loading Issued MR Report inputs: {e}")