# tabs/dashboard_report_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, re, json
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

from fpdf import FPDF
from PIL import Image, ImageDraw, ImageFont # <-- Added PIL imports
from utils import resource_path
from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry
import config

class DashboardReportTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        super().__init__(parent, app_instance, automation_key="dashboard_report")
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) # Main notebook takes up space
        
        # --- FIX: Use correct headers from the final dashboard detail page ---
        self.report_headers = [
            "S No.", "District", "Block", "GP", "Implementing Agency", 
            "Project Name with code", "E-MR No.", "DateFrom-DateTo"
        ]
        
        self._create_widgets()
        self.load_inputs()

    def _create_widgets(self):
        # Frame for all user input controls
        controls_frame = ctk.CTkFrame(self)
        controls_frame.grid(row=0, column=0, sticky="new", padx=10, pady=10)
        controls_frame.grid_columnconfigure(1, weight=1)

        # --- Input Fields ---
        ctk.CTkLabel(controls_frame, text="State:").grid(row=0, column=0, sticky='w', padx=15, pady=(15, 5))
        self.state_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("dashboard_state"),
                                             app_instance=self.app, history_key="dashboard_state")
        self.state_entry.grid(row=0, column=1, sticky='ew', padx=15, pady=(15, 5))

        ctk.CTkLabel(controls_frame, text="District:").grid(row=1, column=0, sticky='w', padx=15, pady=5)
        self.district_entry = AutocompleteEntry(controls_frame, 
                                                suggestions_list=self.app.history_manager.get_suggestions("dashboard_district"),
                                                app_instance=self.app, history_key="dashboard_district")
        self.district_entry.grid(row=1, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Block:").grid(row=2, column=0, sticky='w', padx=15, pady=5)
        self.block_entry = AutocompleteEntry(controls_frame, 
                                             suggestions_list=self.app.history_manager.get_suggestions("dashboard_block"),
                                             app_instance=self.app, history_key="dashboard_block")
        self.block_entry.grid(row=2, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(controls_frame, text="Panchayat:").grid(row=3, column=0, sticky='w', padx=15, pady=5)
        self.panchayat_entry = AutocompleteEntry(controls_frame, 
                                                 suggestions_list=self.app.history_manager.get_suggestions("dashboard_panchayat"),
                                                 app_instance=self.app, history_key="dashboard_panchayat")
        self.panchayat_entry.grid(row=3, column=1, sticky='ew', padx=15, pady=5)

        # --- UPDATE: Delay Column Dropdown Options ---
        ctk.CTkLabel(controls_frame, text="Delay Column:").grid(row=4, column=0, sticky='w', padx=15, pady=5)
        self.delay_column_options = [
            "Attendance not filled in T+2 days",
            "Measurement Book not filled in T+5 days",
            "Wagelist not Sent in T+6 days",
            "Pending for I sig FTO in T+7 days",
            "Pending for II sig FTO in T+8 days"
            # Add back other options if needed from the previous version
        ]
        self.delay_column_entry = ctk.CTkComboBox(controls_frame, values=self.delay_column_options)
        self.delay_column_entry.grid(row=4, column=1, sticky='ew', padx=15, pady=5)
        # Default to the first option
        if self.delay_column_options:
            self.delay_column_entry.set(self.delay_column_options[0])

        action_frame = self._create_action_buttons(parent_frame=controls_frame)
        action_frame.grid(row=5, column=0, columnspan=2, pady=10)

        # --- Output Tabs ---
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        workcode_tab = notebook.add("Workcode List")
        results_tab = notebook.add("Results Table")
        self._create_log_and_status_area(parent_notebook=notebook)

        # 1. Workcode List Tab
        workcode_tab.grid_columnconfigure(0, weight=1)
        workcode_tab.grid_rowconfigure(1, weight=1)
        
        copy_frame = ctk.CTkFrame(workcode_tab, fg_color="transparent")
        copy_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        self.copy_wc_button = ctk.CTkButton(copy_frame, text="Copy Workcodes", command=self._copy_workcodes)
        self.copy_wc_button.pack(side="left")

        # --- NEW BUTTON ---
        self.run_mr_fill_button = ctk.CTkButton(copy_frame,
                                                  text="Run MR Fill",
                                                  command=self._run_mr_fill,
                                                  fg_color="#4A55A2", # Purplish-blue
                                                  hover_color="#5E69B8")
        self.run_mr_fill_button.pack_forget() # Hide it initially
        # --- END NEW BUTTON ---

        self.workcode_textbox = ctk.CTkTextbox(workcode_tab, state="disabled")
        self.workcode_textbox.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        # 2. Results Tab (Table)
        results_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)
        
        export_frame = ctk.CTkFrame(results_tab, fg_color="transparent")
        export_frame.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.export_button = ctk.CTkButton(export_frame, text="Export Report", command=self.export_report)
        self.export_button.pack(side="left")
        
        # --- UPDATE: Added PNG to export options ---
        self.export_format_menu = ctk.CTkOptionMenu(export_frame, values=["Excel (.xlsx)", "PDF (.pdf)", "PNG (.png)"])
        # --- END UPDATE ---
        
        self.export_format_menu.pack(side="left", padx=5)

        # --- FIX: Update Treeview columns and widths ---
        self.results_tree = ttk.Treeview(results_tab, columns=self.report_headers, show='headings')
        for col in self.report_headers: 
            self.results_tree.heading(col, text=col)
            
        # Adjust column widths based on new headers
        self.results_tree.column("S No.", width=40, anchor='center')
        self.results_tree.column("District", width=100)
        self.results_tree.column("Block", width=100)
        self.results_tree.column("GP", width=100)
        self.results_tree.column("Implementing Agency", width=150)
        self.results_tree.column("Project Name with code", width=350)
        self.results_tree.column("E-MR No.", width=80)
        self.results_tree.column("DateFrom-DateTo", width=150)
        # --- END FIX ---

        self.results_tree.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree)

    # --- Methods (set_ui_state, reset_ui, start_automation, _solve_captcha are unchanged) ---
    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        self.state_entry.configure(state=state)
        self.district_entry.configure(state=state)
        self.block_entry.configure(state=state)
        self.panchayat_entry.configure(state=state)
        self.delay_column_entry.configure(state=state)
        self.run_mr_fill_button.configure(state=state)

    def reset_ui(self):
        pass
        
    def start_automation(self):
        self.run_mr_fill_button.pack_forget()
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        self._update_workcode_textbox("") # Clear workcode list
        
        inputs = {
            'state': self.state_entry.get().strip(), 
            'district': self.district_entry.get().strip(), 
            'block': self.block_entry.get().strip(),
            'panchayat': self.panchayat_entry.get().strip(),
            'delay_column': self.delay_column_entry.get().strip()
        }
        
        if not all([inputs['state'], inputs['district'], inputs['block'], inputs['panchayat'], inputs['delay_column']]):
            messagebox.showwarning("Input Error", "All fields are required."); return
        
        self.save_inputs(inputs)
        self.app.update_history("dashboard_state", inputs['state'])
        self.app.update_history("dashboard_district", inputs['district'])
        self.app.update_history("dashboard_block", inputs['block'])
        self.app.update_history("dashboard_panchayat", inputs['panchayat'])
        
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(inputs,))

    def _solve_captcha(self, driver, wait):
        """Solves the math CAPTCHA on the MIS report page."""
        self.app.log_message(self.log_display, "Attempting to solve CAPTCHA...")
        captcha_label_id = "ContentPlaceHolder1_lblStopSpam"; captcha_textbox_id = "ContentPlaceHolder1_txtCaptcha"; verify_button_id = "ContentPlaceHolder1_btnLogin"
        try:
            captcha_element = wait.until(EC.presence_of_element_located((By.ID, captcha_label_id)))
            captcha_text = captcha_element.text
            match = re.search(r'(\d+)\s*([+\-*])\s*(\d+)', captcha_text)
            if not match: raise ValueError("Could not parse CAPTCHA expression.")
            num1, operator, num2 = match.groups(); num1, num2 = int(num1), int(num2)
            result = 0
            if operator == '+': result = num1 + num2
            elif operator == '-': result = num1 - num2
            elif operator == '*': result = num1 * num2
            self.app.log_message(self.log_display, f"Solved: {captcha_text.strip()} = {result}")
            driver.find_element(By.ID, captcha_textbox_id).send_keys(str(result))
            driver.find_element(By.ID, verify_button_id).click()
            # Wait briefly to see if an error message appears
            time.sleep(1)
            if "Invalid Captcha Code" in driver.page_source:
                raise ValueError("CAPTCHA verification failed.")
            return True
        except TimeoutException:
            # If CAPTCHA elements are not found, assume it's already solved or not present
            self.app.log_message(self.log_display, "CAPTCHA not found or already bypassed.", "info")
            return True # Continue as if successful
        except ValueError as e:
            self.app.log_message(self.log_display, f"CAPTCHA Error: {e}", "error")
            raise # Re-raise the error to be caught by the main logic

    def run_automation_logic(self, inputs, retries=1):
        # --- Set Initial Status ---
        self.app.after(0, self.set_ui_state, True)
        self.app.after(0, self.app.set_status, "Starting Dashboard Report...") # App-wide status
        self.app.after(0, self.update_status, "Initializing...", 0.0) # Tab-specific status
        self.app.clear_log(self.log_display)
        self.app.log_message(self.log_display, "Starting Dashboard Report automation...")

        try:
            driver = self.app.get_driver()
            if not driver:
                self.app.after(0, self.app.set_status, "Browser not found")
                return # Exit early

            wait = WebDriverWait(driver, 20)

            # --- Status Update ---
            self.app.after(0, self.app.set_status, "Navigating to MIS portal...")
            self.app.after(0, self.update_status, "Navigating...", 0.05)
            self.app.log_message(self.log_display, "Navigating to MIS portal...")
            driver.get(config.MIS_REPORTS_CONFIG["base_url"])

            # --- Status Update ---
            self.app.after(0, self.app.set_status, "Solving CAPTCHA...")
            self.app.after(0, self.update_status, "Solving CAPTCHA...", 0.1)
            self._solve_captcha(driver, wait) # Handles potential failure
            self.app.log_message(self.log_display, "CAPTCHA step passed. Selecting state...")

            # --- Status Update ---
            self.app.after(0, self.app.set_status, f"Selecting State: {inputs['state']}...")
            self.app.after(0, self.update_status, "Selecting State...", 0.15)
            state_select = wait.until(EC.element_to_be_clickable((By.ID, "ContentPlaceHolder1_ddl_States")))
            Select(state_select).select_by_visible_text(inputs['state'].upper())
            wait.until(EC.presence_of_element_located((By.LINK_TEXT, "Dashboard for Delay Monitoring System")))

            # --- Status Update ---
            self.app.after(0, self.app.set_status, "Opening Dashboard Report...")
            self.app.after(0, self.update_status, "Opening Dashboard...", 0.2)
            self.app.log_message(self.log_display, "Clicking 'Dashboard for Delay Monitoring System'...")
            report_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Dashboard for Delay Monitoring System")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", report_link)
            time.sleep(1); report_link.click()

            # Optional state re-selection
            # self.app.log_message(self.log_display, "Selecting State again (if required)...")
            # wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['state'].upper()))).click()

            # --- Status Update ---
            self.app.after(0, self.app.set_status, f"Selecting District: {inputs['district']}...")
            self.app.after(0, self.update_status, "Selecting District...", 0.25)
            self.app.log_message(self.log_display, f"Drilling down to District: {inputs['district']}")
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['district'].upper()))).click()

            # --- Status Update ---
            self.app.after(0, self.app.set_status, f"Selecting Block: {inputs['block']}...")
            self.app.after(0, self.update_status, "Selecting Block...", 0.3)
            self.app.log_message(self.log_display, f"Drilling down to Block: {inputs['block']}")
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['block'].upper()))).click()

            # --- Status Update ---
            self.app.after(0, self.app.set_status, f"Finding Panchayat: {inputs['panchayat']}...")
            self.app.after(0, self.update_status, "Finding Panchayat...", 0.35)
            self.app.log_message(self.log_display, f"Finding Panchayat row: {inputs['panchayat']}")
            main_table_xpath = "//table[.//b[text()='S No.'] and .//b[text()='Panchayat']]"
            wait.until(EC.presence_of_element_located((By.XPATH, f"{main_table_xpath}//tr[1]/td/b[text()='Panchayat']")))

            panchayat_row_xpath = f"{main_table_xpath}//tr[td[2][normalize-space()='{inputs['panchayat']}']]"
            panchayat_row = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, panchayat_row_xpath)))

            # --- Status Update ---
            self.app.after(0, self.app.set_status, f"Finding Column: {inputs['delay_column']}...")
            self.app.after(0, self.update_status, "Finding Column...", 0.4)
            self.app.log_message(self.log_display, f"Finding column header: {inputs['delay_column']}")
            header_cells = driver.find_elements(By.XPATH, f"{main_table_xpath}//tr[.//b[contains(text(), 'T+2')]]/td/b")

            target_col_index = -1
            for i, th_b in enumerate(header_cells):
                header_text = ' '.join(th_b.text.split()).lower().strip()
                search_text = ' '.join(inputs['delay_column'].split()).lower().strip()
                if search_text == header_text:
                    target_col_index = i + 2
                    break

            if target_col_index == -1:
                raise ValueError(f"Could not find column header matching '{inputs['delay_column']}'")

            # --- Status Update ---
            self.app.after(0, self.app.set_status, "Clicking Report Link...")
            self.app.after(0, self.update_status, "Clicking Link...", 0.45)
            self.app.log_message(self.log_display, f"Found column at index {target_col_index}. Clicking cell link in Panchayat row.")
            row_cells = panchayat_row.find_elements(By.TAG_NAME, "td")

            if target_col_index >= len(row_cells):
                 raise IndexError(f"Calculated column index {target_col_index} is out of bounds for the row.")

            target_cell = row_cells[target_col_index]

            try:
                target_link = target_cell.find_element(By.TAG_NAME, "a")
                # We don't care about the text, just click it.
                self.app.log_message(self.log_display, f"Found link (text: '{target_link.text.strip()}'). Clicking...")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_link)
                time.sleep(0.5)
                target_link.click()

            except NoSuchElementException:
                 # This means there is no <a> tag in the cell.
                 cell_text = target_cell.text.strip()
                 if cell_text == '0':
                    # This is the *only* place we should check for '0' (when it's not a link).
                    self.app.log_message(self.log_display, f"Column '{inputs['delay_column']}' has value 0 (not a link). No data to fetch.", "warning")
                    messagebox.showinfo("No Data", f"The selected column '{inputs['delay_column']}' has a value of 0 for {inputs['panchayat']}. No details to display.")
                    self.success_message = None
                    self.app.after(0, self.app.set_status, "No data found")
                    return
                 else:
                    # The cell has text, but it's not '0' and not a link. This is an error.
                    raise ValueError(f"Target cell for column '{inputs['delay_column']}' does not contain a clickable link (text: {cell_text}).")

            # --- Status Update ---
            self.app.after(0, self.app.set_status, "Loading Final Report...")
            self.app.after(0, self.update_status, "Loading Final Report...", 0.5)
            self.app.log_message(self.log_display, "Waiting for final report table...")
            FINAL_TABLE_XPATH = "//table[@bordercolor='green' and .//b[contains(text(), 'E-MR No.')]]"
            table = wait.until(EC.presence_of_element_located((By.XPATH, FINAL_TABLE_XPATH)))
            rows = table.find_elements(By.XPATH, ".//tr[position()>1]") # Skip header row

            total_rows = len(rows)
            if total_rows == 0:
                self.app.log_message(self.log_display, "Final report table is empty.", "warning")
                messagebox.showinfo("No Data", f"No detailed records found for {inputs['panchayat']} under '{inputs['delay_column']}'.")
                self.success_message = None
                 # --- Status Update on early exit ---
                self.app.after(0, self.app.set_status, "No data found")
                return

            self.app.log_message(self.log_display, f"Found {total_rows} records in the final table. Processing...")

            workcode_list = []
            pending_mr_count = 0

            for i, row in enumerate(rows):
                if self.app.stop_events[self.automation_key].is_set():
                    self.app.log_message(self.log_display, "Stop signal received.", "warning")
                    break # Exit loop if stopped

                # --- Status Update (Inside Loop) ---
                # Calculate progress from 0.5 to 0.95 based on row processing
                progress = 0.5 + ( (i + 1) / total_rows ) * 0.45
                status_msg = f"Processing row {i+1}/{total_rows}"
                self.app.after(0, self.app.set_status, status_msg)
                self.app.after(0, self.update_status, status_msg, progress)
                # ---

                cells = row.find_elements(By.TAG_NAME, "td")
                if not cells or len(cells) < len(self.report_headers):
                    self.app.log_message(self.log_display, f"Skipping row {i+1}, expected at least {len(self.report_headers)} columns, found {len(cells)}.", "warning")
                    continue

                scraped_data = [cell.text.strip() for cell in cells[:len(self.report_headers)]]
                project_name_code = scraped_data[5]
                work_code_match = re.search(r'\(([^)]+)\)$', project_name_code)
                work_code = work_code_match.group(1).strip() if work_code_match else "N/A"
                pending_mr_count += 1
                row_data = tuple(scraped_data)

                self.app.after(0, lambda data=row_data: self.results_tree.insert("", "end", values=data))
                if work_code != "N/A":
                    workcode_list.append(work_code)

            # Check if stopped *after* the loop
            if self.app.stop_events[self.automation_key].is_set():
                 self.app.log_message(self.log_display, "Automation stopped by user.", "warning")
                 self.success_message = None # Prevent success message
                 return # Exit cleanly


            # Update workcode list
            # unique_workcodes = list(dict.fromkeys(workcode_list)) # <-- DONT remove duplicates
            self.app.after(0, self._update_workcode_textbox, "\n".join(workcode_list)) # <-- Use the full list

            self.app.log_message(self.log_display, f"Processing complete. Found {pending_mr_count} pending MRs listed.", "success")
            self.success_message = f"Dashboard Report automation has finished.\n{pending_mr_count} Pending MR to Fill in {inputs['panchayat']}."

        except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
            if "Session Expired" in driver.page_source and retries > 0:
                self.app.log_message(self.log_display, "Session expired, attempting retry...", "warning")
                # --- Status Update ---
                self.app.after(0, self.app.set_status, "Session expired, retrying...")
                self.app.after(0, self.update_status, "Retrying...", 0.0)
                self.run_automation_logic(inputs, retries - 1)
                return # Stop current execution after scheduling retry
            error_msg = f"A browser error occurred: {str(e).splitlines()[0]}"
            self.app.log_message(self.log_display, error_msg, "error")
            messagebox.showerror("Automation Error", error_msg)
            # --- Status Update ---
            self.app.after(0, self.app.set_status, "Browser Error")
            self.success_message = None
        except ValueError as e: # Catch CAPTCHA errors, column find errors, etc.
             error_msg = f"Data processing error: {e}"
             self.app.log_message(self.log_display, error_msg, "error")
             messagebox.showerror("Automation Error", error_msg)
             # --- Status Update ---
             self.app.after(0, self.app.set_status, "Data Error")
             self.success_message = None
        except Exception as e:
            self.app.log_message(self.log_display, f"An unexpected error occurred: {e}", "error")
            messagebox.showerror("Critical Error", f"An unexpected error occurred: {e}")
             # --- Status Update ---
            self.app.after(0, self.app.set_status, "Unexpected Error")
            self.success_message = None
        finally:
            # --- Final Status Updates ---
            # Ensure UI is re-enabled even if an error occurred
            self.app.after(0, self.set_ui_state, False)
            
            # Set final status based on whether it was stopped or completed/failed
            final_app_status = "Automation Stopped" if self.app.stop_events[self.automation_key].is_set() else \
                              ("Automation Finished" if hasattr(self, 'success_message') and self.success_message else "Automation Failed")
            final_tab_status = "Stopped" if self.app.stop_events[self.automation_key].is_set() else \
                              ("Finished" if hasattr(self, 'success_message') and self.success_message else "Failed")

            self.app.after(0, self.app.set_status, final_app_status)
            self.app.after(0, self.update_status, final_tab_status, 1.0)

            # Reset to 'Ready' after a delay if not stopped
            if not self.app.stop_events[self.automation_key].is_set():
                 self.app.after(5000, lambda: self.app.set_status("Ready")) # Reset app status
                 self.app.after(5000, lambda: self.update_status("Ready", 0.0)) # Reset tab status

            # Only show success message if it exists and wasn't stopped
            if hasattr(self, 'success_message') and self.success_message and not self.app.stop_events[self.automation_key].is_set():
                self.app.after(100, lambda: messagebox.showinfo("Complete", self.success_message))
            # ---
                # --- NEW: Show MR Fill button if conditions are met ---
                if inputs['delay_column'] == "Attendance not filled in T+2 days":
                    self.app.after(0, lambda: self.run_mr_fill_button.pack(side="left", padx=(10, 0)))
                # --- END NEW ---

    def _update_workcode_textbox(self, text):
        self.workcode_textbox.configure(state="normal")
        self.workcode_textbox.delete("1.0", tkinter.END)
        self.workcode_textbox.insert("1.0", text)
        self.workcode_textbox.configure(state="disabled")

    def _copy_workcodes(self):
        text = self.workcode_textbox.get("1.0", tkinter.END).strip()
        if text:
            self.app.clipboard_clear()
            self.app.clipboard_append(text)
            messagebox.showinfo("Copied", f"{len(text.splitlines())} workcodes copied to clipboard.", parent=self)
        else:
            messagebox.showwarning("Empty", "There are no workcodes to copy.", parent=self)

    def _run_mr_fill(self):
        """Called when the 'Run MR Fill' button is clicked."""
        workcodes = self.workcode_textbox.get("1.0", tkinter.END).strip()
        panchayat_name = self.panchayat_entry.get().strip()

        if not workcodes:
            messagebox.showwarning("No Data", "There are no workcodes to send to the MR Fill tab.", parent=self)
            return
        
        if not panchayat_name:
            messagebox.showwarning("No Data", "Panchayat name is missing. Cannot send to MR Fill tab.", parent=self)
            return

        # Call the new method in the main app
        self.app.switch_to_mr_fill_with_data(workcodes, panchayat_name)
    # --- END NEW METHOD ---

    def export_report(self):
        if not self.results_tree.get_children():
            messagebox.showinfo("No Data", "There are no results to export.")
            return

        panchayat = self.panchayat_entry.get().strip() or "Report"
        safe_panchayat = re.sub(r'[\\/*?:"<>|]', '_', panchayat) 
        
        export_format = self.export_format_menu.get()
        current_year = datetime.now().strftime("%Y")
        current_date_str = datetime.now().strftime("%d-%b-%Y") # e.g., 29-Oct-2025

        headers = self.report_headers
        data = [self.results_tree.item(item, 'values') for item in self.results_tree.get_children()]

        # --- FIX: Define Correct PDF Title ---
        pdf_title = f"Pending for Filling MusterRoll - {panchayat}" 
        date_str_header = f"Date - {datetime.now().strftime('%d-%m-%Y')}" # For PDF header
        excel_title = f"{pdf_title} {date_str_header}" # Combine for Excel
        # --- END FIX ---
        
        # ... (Folder creation logic remains the same) ...
        downloads_path = self.app.get_user_downloads_path()
        target_dir = os.path.join(downloads_path, f"Reports {current_year}", safe_panchayat)
        try:
            os.makedirs(target_dir, exist_ok=True)
        except OSError as e:
             messagebox.showerror("Folder Error", f"Could not create report directory:\n{target_dir}\nError: {e}")
             return

        # --- UPDATE: Handle all export formats (Excel, PDF, PNG) ---
        if "Excel" in export_format:
            ext = ".xlsx"
            file_type_tuple = ("Excel Workbook", "*.xlsx")
            default_filename = f"Pending_MusterRoll_{safe_panchayat}-{current_date_str}{ext}"
        elif "PDF" in export_format:
            ext = ".pdf"
            file_type_tuple = ("PDF Document", "*.pdf")
            default_filename = f"Pending_MusterRoll_{safe_panchayat}-{current_date_str}{ext}"
        elif "PNG" in export_format: # <-- NEW
            ext = ".png"
            file_type_tuple = ("PNG Image", "*.png")
            default_filename = f"Pending_MusterRoll_{safe_panchayat}-{current_date_str}{ext}"
        else:
            return # Should not happen

        file_path = filedialog.asksaveasfilename(
            initialdir=target_dir, 
            initialfile=default_filename, 
            defaultextension=ext,
            filetypes=[file_type_tuple, ("All Files", "*.*")],
            title="Save Report As"
        )
        
        if not file_path: return 

        if "Excel" in export_format:
            # --- Pass the combined title for Excel ---
            success = self._save_to_excel(data, headers, excel_title, file_path) 
            if success:
                messagebox.showinfo("Success", f"Excel report saved successfully to:\n{file_path}")
        
        elif "PDF" in export_format:
            # These widths are for the 8 headers in self.report_headers
            col_widths = [12, 30, 30, 30, 40, 115, 20, 30] 
            total_width_ratio = sum(col_widths)
            effective_page_width = 297 - 20 
            actual_col_widths = [(w / total_width_ratio) * effective_page_width for w in col_widths]
            
            # --- Pass the specific PDF title and date string ---
            success = self.generate_report_pdf(data, headers, actual_col_widths, pdf_title, date_str_header, file_path) 
            if success:
                messagebox.showinfo("Success", f"PDF report saved successfully to:\n{file_path}")

        elif "PNG" in export_format: # <-- NEW
            # Pass the PDF title and date string to the PNG function
            success = self._save_to_png(data, headers, pdf_title, date_str_header, file_path)
            if success:
                messagebox.showinfo("Success", f"PNG report saved successfully to:\n{file_path}")
        # --- END UPDATE ---


    def _save_to_excel(self, data, headers, title, file_path):
        try:
            df = pd.DataFrame(data, columns=headers)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Use a sheet name suitable for Excel (no slashes etc.)
                sheet_name = 'Dashboard Report'
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                worksheet = writer.sheets[sheet_name]
                
                worksheet['A1'] = title
                worksheet['A1'].font = Font(bold=True, size=14)
                worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                # Correct merge cell column count
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers)) 
                
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
                for cell in worksheet["2:2"]:
                    cell.font = header_font
                    cell.fill = header_fill

                for col_idx, col in enumerate(df.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    # Handle potential errors if column data is empty or non-string
                    try:
                        max_length = max(len(str(col)), df[col].astype(str).map(len).max())
                    except (TypeError, ValueError):
                         max_length = len(str(col)) # Fallback to header length
                    adjusted_width = min((max_length + 2), 50) 
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            return True
        except Exception as e:
            messagebox.showerror("Excel Export Error", f"Could not generate Excel report.\nError: {e}", parent=self)
            return False


    def generate_report_pdf(self, data, headers, col_widths, title, date_str, file_path):
        """ Overrides base method to use Unicode font, add footer, adjust formatting, and reduce row gaps. """
        
        class PDFWithFooter(FPDF):
            def footer(self):
                self.set_y(-15) 
                try:
                    # Use regular font style
                    self.set_font(font_name, '', 8) 
                except NameError: 
                    self.set_font('Helvetica', '', 8) 
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
                self.set_xy(self.l_margin, -15)
                self.cell(0, 10, 'Report Generated by NregaBot.com', 0, 0, 'L')

        try:
            pdf = PDFWithFooter(orientation="L", unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=15) # Ensure auto page break with margin
            pdf.add_page()
            
            try:
                font_path_regular = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
                font_path_bold = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
                pdf.add_font("NotoSansDevanagari", "", font_path_regular, uni=True)
                pdf.add_font("NotoSansDevanagari", "B", font_path_bold, uni=True)
                font_name = "NotoSansDevanagari"
            except RuntimeError:
                font_name = "Helvetica" 

            # --- Title ---
            pdf.set_font(font_name, "B", 14) 
            pdf.cell(0, 10, title, 0, 1, "C") # Title is passed in correctly now
            pdf.set_font(font_name, "", 10) 
            pdf.cell(0, 8, date_str, 0, 1, "R") 
            pdf.ln(4) 

            # --- Headers ---
            pdf.set_font(font_name, "B", 7) 
            pdf.set_fill_color(200, 220, 255)
            header_height = 8 # Define header height
            for i, header in enumerate(headers):
                if i < len(col_widths):
                    pdf.cell(col_widths[i], header_height, header, 1, 0, "C", fill=True) 
                else:
                    pdf.cell(10, header_height, header, 1, 0, "C", fill=True) 
            pdf.ln()

            # --- Data Rows ---
            pdf.set_font(font_name, "", 6) 
            line_height = 4 # Define line height
            
            for row_data in data:
                if len(row_data) != len(headers):
                    continue

                # --- Calculate row height based on content wrapping ---
                max_lines = 1
                for i, cell_text in enumerate(row_data):
                    if i < len(col_widths): 
                        lines = pdf.multi_cell(col_widths[i], line_height, str(cell_text), border=0, align='L', split_only=True)
                        # Ensure at least 1 line is counted even for empty strings
                        current_lines = len(lines) if lines else 1 
                        if current_lines > max_lines: max_lines = current_lines
                
                row_height = line_height * max_lines
                
                # --- Check for page break BEFORE drawing the row ---
                if pdf.get_y() + row_height > pdf.page_break_trigger:
                    pdf.add_page()
                    # Redraw headers on new page
                    pdf.set_font(font_name, "B", 7)
                    for i, header in enumerate(headers):
                        if i < len(col_widths):
                             pdf.cell(col_widths[i], header_height, header, 1, 0, "C", fill=True)
                        else:
                             pdf.cell(10, header_height, header, 1, 0, "C", fill=True)
                    pdf.ln()
                    pdf.set_font(font_name, "", 6) # Reset data font

                # --- Draw the row using multi_cell for wrapping ---
                y_start = pdf.get_y()
                x_start = pdf.l_margin 
                
                for i, cell_text in enumerate(row_data):
                    if i < len(col_widths): 
                        col_width = col_widths[i]
                        x_current = x_start + sum(col_widths[:i]) 
                        pdf.set_xy(x_current, y_start) 
                        # --- Draw the cell with calculated height ---
                        pdf.multi_cell(col_width, line_height, str(cell_text), border=1, align='L', max_line_height=line_height) 
                
                # --- FIX: Set Y position explicitly based on calculated row height ---
                pdf.set_y(y_start + row_height) 
                # --- END FIX ---

            pdf.output(file_path)
            return True
        except Exception as e:
            messagebox.showerror("PDF Export Error", f"Could not generate PDF report.\nError: {e}", parent=self)
            return False

    # --- NEW METHOD: _save_to_png ---
    def _save_to_png(self, data, headers, title, date_str, file_path):
        """Generates a professional-looking report as a PNG image."""
        try:
            # --- Font setup ---
            try:
                font_path_regular = resource_path("assets/fonts/NotoSansDevanagari-Regular.ttf")
                font_path_bold = resource_path("assets/fonts/NotoSansDevanagari-Bold.ttf")
                font_title = ImageFont.truetype(font_path_bold, 28)
                font_date = ImageFont.truetype(font_path_regular, 18)
                font_header = ImageFont.truetype(font_path_bold, 16)
                font_body = ImageFont.truetype(font_path_regular, 14)
            except IOError:
                self.app.log_message(self.log_display, "Warning: NotoSansDevanagari fonts not found. Using default PIL fonts. Ensure 'assets/fonts' exist.", "warning")
                font_title = ImageFont.load_default(size=28)
                font_date = ImageFont.load_default(size=18)
                font_header = ImageFont.load_default(size=16)
                font_body = ImageFont.load_default(size=14)

            # --- Image dimensions and colors ---
            img_width = 2400  # High resolution for readability
            margin_x = 80
            margin_y = 60
            
            header_bg_color = (220, 235, 255) # Light blue
            row_even_bg_color = (255, 255, 255) # White
            row_odd_bg_color = (245, 245, 245)  # Light grey
            text_color = (0, 0, 0) # Black
            border_color = (180, 180, 180) # Grey

            draw_start_y = margin_y # Y position to start drawing content

            # --- Calculate column widths (heuristic approach) ---
            # Based on the 8 headers for this report
            # "S No.", "District", "Block", "GP", "Implementing Agency", "Project Name with code", "E-MR No.", "DateFrom-DateTo"
            # Give "Project Name" the most space.
            base_col_widths = [0.05, 0.10, 0.10, 0.10, 0.15, 0.35, 0.07, 0.08]
            
            available_width = img_width - (2 * margin_x)
            col_widths_pixels = [w * available_width for w in base_col_widths]
            
            # --- Auto-adjust column widths based on content ---
            # This is a simplified version; a full auto-fit is complex.
            # We'll stick to ratio but ensure headers fit.
            for i, header in enumerate(headers):
                header_width = font_header.getlength(header) + 40 # Add padding
                if col_widths_pixels[i] < header_width:
                    # If header is wider, steal space from the widest column (Project Name)
                    diff = header_width - col_widths_pixels[i]
                    col_widths_pixels[i] = header_width
                    col_widths_pixels[5] -= diff # Index 5 is "Project Name"
            
            # Ensure no column is negative
            if col_widths_pixels[5] < 100: col_widths_pixels[5] = 100
            
            # Re-normalize to fit available width exactly
            current_total_width = sum(col_widths_pixels)
            scale_factor = available_width / current_total_width
            col_widths_pixels = [w * scale_factor for w in col_widths_pixels]

            # Estimate total height required
            title_bbox = font_title.getbbox(title)
            date_bbox = font_date.getbbox(date_str)
            title_height = title_bbox[3] - title_bbox[1]
            date_height = date_bbox[3] - date_bbox[1]

            # Start with a base image, will expand if needed
            initial_height = 1600 # Start with a reasonable height
            img = Image.new("RGB", (img_width, initial_height), (255, 255, 255))
            draw = ImageDraw.Draw(img)

            current_y = margin_y
            
            # --- Draw Title ---
            title_text_width = font_title.getlength(title)
            title_x = (img_width - title_text_width) / 2
            draw.text((title_x, current_y), title, font=font_title, fill=text_color)
            current_y += title_height + 5

            # --- Draw Date ---
            date_text_width = font_date.getlength(date_str)
            date_x = img_width - margin_x - date_text_width
            draw.text((date_x, current_y), date_str, font=font_date, fill=text_color)
            current_y += date_height + 20 # Space after date and before table

            # --- Draw Headers ---
            header_y_start = current_y
            header_height = 0
            # Calculate max header height considering wrapping
            for i, header in enumerate(headers):
                wrapped_header = self._wrap_text(header, font_header, col_widths_pixels[i] - 10)
                line_height = font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1] # Basic line height
                header_height = max(header_height, len(wrapped_header) * line_height + 10) # Add padding
            
            # Draw header cells
            current_x = margin_x
            for i, header in enumerate(headers):
                draw.rectangle([current_x, header_y_start, current_x + col_widths_pixels[i], header_y_start + header_height], fill=header_bg_color, outline=border_color, width=1)
                
                wrapped_header = self._wrap_text(header, font_header, col_widths_pixels[i] - 20)
                line_height = font_header.getbbox("Tg")[3] - font_header.getbbox("Tg")[1]
                total_text_height = len(wrapped_header) * line_height
                text_y = header_y_start + (header_height - total_text_height) / 2 # Center vertically
                
                for line in wrapped_header:
                    line_width = font_header.getlength(line)
                    draw.text((current_x + (col_widths_pixels[i] - line_width) / 2, text_y), line, font=font_header, fill=text_color)
                    text_y += line_height # Move to next line
                current_x += col_widths_pixels[i]
            current_y += header_height

            # --- Draw Data Rows ---
            for row_idx, row_data in enumerate(data):
                row_bg_color = row_even_bg_color if row_idx % 2 == 0 else row_odd_bg_color

                # Calculate row height dynamically based on content wrapping
                max_row_text_height = 0
                temp_wrapped_cells = []
                for i, cell_text in enumerate(row_data):
                    wrapped_lines = self._wrap_text(str(cell_text), font_body, col_widths_pixels[i] - 20) # 20 for internal padding
                    temp_wrapped_cells.append(wrapped_lines)
                    line_height = font_body.getbbox("Tg")[3] - font_body.getbbox("Tg")[1] # Basic line height
                    max_row_text_height = max(max_row_text_height, len(wrapped_lines) * line_height)

                row_data_height = max_row_text_height + 10 # Add vertical padding
                line_height = font_body.getbbox("Tg")[3] - font_body.getbbox("Tg")[1]

                # If current_y plus next row height exceeds image height, expand image
                if current_y + row_data_height + margin_y > img.height:
                    new_height = int(img.height + (row_data_height + margin_y) * 10) # Expand by 10 rows
                    new_img = Image.new("RGB", (img_width, new_height), (255, 255, 255))
                    new_img.paste(img, (0, 0))
                    img = new_img
                    draw = ImageDraw.Draw(img) # Update draw object for new image

                current_x = margin_x
                for i, cell_text in enumerate(row_data):
                    # Draw cell background
                    draw.rectangle([current_x, current_y, current_x + col_widths_pixels[i], current_y + row_data_height], fill=row_bg_color, outline=border_color, width=1)
                    
                    # Draw wrapped text
                    wrapped_lines = temp_wrapped_cells[i]
                    text_y = current_y + 5 # Small top padding
                    for line in wrapped_lines:
                        draw.text((current_x + 10, text_y), line, font=font_body, fill=text_color) # 10 for left padding
                        text_y += line_height # Move to next line
                    current_x += col_widths_pixels[i]
                current_y += row_data_height

            # --- Draw Footer ---
            current_y += 15 # Space after table
            footer_text = "Report Generated by NregaBot.com"
            footer_font = font_body
            footer_bbox = footer_font.getbbox(footer_text)
            footer_height = footer_bbox[3] - footer_bbox[1]
            footer_y_pos = current_y + 10

            # Check if image needs expansion for footer + bottom margin
            if footer_y_pos + footer_height + margin_y > img.height:
                new_height = int(footer_y_pos + footer_height + margin_y)
                new_img = Image.new("RGB", (img_width, new_height), (255, 255, 255))
                new_img.paste(img, (0, 0))
                img = new_img
                draw = ImageDraw.Draw(img)
            
            # Draw footer text
            draw.text((margin_x, footer_y_pos), footer_text, font=footer_font, fill=text_color)
            current_y = footer_y_pos + footer_height # Update final Y position
            # --- End Footer ---

            # Crop image to actual content
            final_img = img.crop((0, 0, img_width, current_y + margin_y))
            final_img.save(file_path, "PNG", dpi=(300, 300)) # Save at 300 DPI for quality
            return True
        except Exception as e:
            messagebox.showerror("PNG Export Error", f"Could not generate PNG report.\nError: {e}", parent=self)
            return False

    # --- NEW METHOD: _wrap_text ---
    def _wrap_text(self, text, font, max_width):
        """Helper to wrap text for Pillow."""
        if not text:
            return [""]
        words = text.split(' ')
        lines = []
        current_line = []
        for word in words:
            if font.getlength(' '.join(current_line + [word])) <= max_width:
                current_line.append(word)
            else:
                if current_line: # Handle case where a single word is too long
                    lines.append(' '.join(current_line))
                current_line = [word]
                # Handle single word longer than max_width
                while font.getlength(' '.join(current_line)) > max_width:
                    # This is a simple character-based break, not ideal but works
                    part = ""
                    for i, char in enumerate(current_line[0]):
                        if font.getlength(part + char) > max_width:
                            lines.append(part)
                            current_line[0] = current_line[0][i:]
                            part = ""
                            break
                        part += char
                    else:
                        break # Word fits now
        
        if current_line:
            lines.append(' '.join(current_line))
            
        return lines if lines else [""]

        
    def save_inputs(self, inputs):
        """Saves non-sensitive inputs for this tab."""
        save_data = {k: inputs.get(k) for k in ('state', 'district', 'block', 'panchayat')}
        try:
            config_file = self.app.get_data_path("dashboard_report_inputs.json")
            with open(config_file, 'w') as f:
                json.dump(save_data, f, indent=4)
        except Exception as e:
            print(f"Error saving Dashboard Report inputs: {e}")

    def load_inputs(self):
        """Loads saved inputs for this tab."""
        try:
            config_file = self.app.get_data_path("dashboard_report_inputs.json")
            if not os.path.exists(config_file): return
            
            with open(config_file, 'r') as f: data = json.load(f)
            
            self.state_entry.delete(0, 'end') # Clear before insert
            self.state_entry.insert(0, data.get('state', ''))
            self.district_entry.delete(0, 'end')
            self.district_entry.insert(0, data.get('district', ''))
            self.block_entry.delete(0, 'end')
            self.block_entry.insert(0, data.get('block', ''))
            self.panchayat_entry.delete(0, 'end')
            self.panchayat_entry.insert(0, data.get('panchayat', ''))
             # Don't load delay column, let it default
        except Exception as e:
            print(f"Error loading Dashboard Report inputs: {e}")

            