# tabs/mis_reports_tab.py
import tkinter
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import time, os, json
from datetime import datetime
import pandas as pd
import re

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

# --- MODIFIED IMPORTS ---
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .base_tab import BaseAutomationTab
from .autocomplete_widget import AutocompleteEntry

class MisReportsTab(BaseAutomationTab):
    def __init__(self, parent, app_instance):
        super().__init__(parent, app_instance, automation_key="mis_reports")
        self.config_file = self.app.get_data_path("mis_reports_inputs.json")
        self.report_checkboxes = {}
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        self._create_widgets()
        self.load_inputs()

    def _create_widgets(self):
        # Main tab view to organize Settings, Results, and Logs
        notebook = ctk.CTkTabview(self)
        notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        settings_tab = notebook.add("Settings")
        results_tab = notebook.add("Results")
        self._create_log_and_status_area(notebook) # Creates "Logs & Status" tab

        # Configure the layout for the tabs
        settings_tab.grid_rowconfigure(1, weight=1)
        settings_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)
        results_tab.grid_columnconfigure(0, weight=1)
        
        # 1. Populate the "Settings" Tab
        settings_container = ctk.CTkFrame(settings_tab, fg_color="transparent")
        settings_container.grid(row=0, column=0, sticky="nsew", padx=5)
        settings_container.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(settings_container, text="State:").grid(row=0, column=0, sticky='w', padx=15, pady=5)
        self.state_entry = AutocompleteEntry(settings_container, suggestions_list=self.app.history_manager.get_suggestions("mis_state"))
        self.state_entry.grid(row=0, column=1, sticky='ew', padx=15, pady=5)

        ctk.CTkLabel(settings_container, text="District:").grid(row=1, column=0, sticky='w', padx=15, pady=5)
        self.district_entry = AutocompleteEntry(settings_container, suggestions_list=self.app.history_manager.get_suggestions("mis_district"))
        self.district_entry.grid(row=1, column=1, sticky='ew', padx=15, pady=5)
        
        ctk.CTkLabel(settings_container, text="Block:").grid(row=2, column=0, sticky='w', padx=15, pady=5)
        self.block_entry = AutocompleteEntry(settings_container, suggestions_list=self.app.history_manager.get_suggestions("mis_block"))
        self.block_entry.grid(row=2, column=1, sticky='ew', padx=15, pady=5)

        # Checkbox list for reports
        reports_frame = ctk.CTkFrame(settings_tab)
        reports_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=10)
        reports_frame.grid_columnconfigure(0, weight=1)
        reports_frame.grid_rowconfigure(1, weight=1)
        
        reports_header = ctk.CTkFrame(reports_frame, fg_color="transparent")
        reports_header.grid(row=0, column=0, sticky='ew', padx=10, pady=(5,0))
        
        ctk.CTkLabel(reports_header, text="Reports to Download:", font=ctk.CTkFont(weight="bold")).pack(side="left")
        
        btn_frame = ctk.CTkFrame(reports_header, fg_color="transparent")
        btn_frame.pack(side="right")
        ctk.CTkButton(btn_frame, text="Select All", width=100, command=self._toggle_all_checkboxes).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="Deselect All", width=100, command=lambda: self._toggle_all_checkboxes(select=False)).pack(side="left")
        
        scrollable_frame = ctk.CTkScrollableFrame(reports_frame, label_text="")
        scrollable_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=5)

        self.report_list = [
            "Dashboard for Delay Monitoring System", "MGNREGS daily status as per e-muster issued",
            "Employment Pattern During the year", "SC ST Employment Status",
            "Aadhaar Status with NPCI Mapper", "Employment Provided Period wise",
            "Yearly Work Completion Rate", "Rejected Wage/Material/admin Transaction Reconciliation",
            "Verification of Job cards"
        ]
        
        self.report_checkboxes = {}
        for report_name in self.report_list:
            var = tkinter.IntVar(value=1)
            cb = ctk.CTkCheckBox(scrollable_frame, text=report_name, variable=var)
            cb.pack(anchor="w", padx=10, pady=5)
            self.report_checkboxes[report_name] = var

        action_frame = self._create_action_buttons(parent_frame=settings_tab)
        action_frame.grid(row=2, column=0, pady=10)

        # 2. Populate the "Results" Tab
        cols = ("Report Name", "Status", "Details")
        self.results_tree = ttk.Treeview(results_tab, columns=cols, show='headings')
        for col in cols: self.results_tree.heading(col, text=col)
        self.results_tree.column("Report Name", width=300); self.results_tree.column("Status", width=100); self.results_tree.column("Details", width=300)
        self.results_tree.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        scrollbar = ctk.CTkScrollbar(results_tab, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set); scrollbar.grid(row=1, column=1, sticky='ns')
        self.style_treeview(self.results_tree)

    def _toggle_all_checkboxes(self, select=True):
        for var in self.report_checkboxes.values():
            var.set(1 if select else 0)
    
    def set_ui_state(self, running: bool):
        self.set_common_ui_state(running)
        state = "disabled" if running else "normal"
        self.state_entry.configure(state=state)
        self.district_entry.configure(state=state)
        self.block_entry.configure(state=state)
        for cb in self.winfo_descendants():
            if isinstance(cb, ctk.CTkCheckBox) or isinstance(cb, ctk.CTkButton):
                if cb != self.stop_button:
                    cb.configure(state=state)

    def start_automation(self):
        for item in self.results_tree.get_children(): self.results_tree.delete(item)
        
        selected_reports = [name for name, var in self.report_checkboxes.items() if var.get() == 1]
        
        inputs = {'state': self.state_entry.get().strip(), 'district': self.district_entry.get().strip(), 'block': self.block_entry.get().strip(), 'reports': selected_reports}
        
        if not all([inputs['state'], inputs['district'], inputs['block'], inputs['reports']]):
            messagebox.showwarning("Input Error", "State, District, Block, and at least one Report are required."); return
        
        self.save_inputs({'state': inputs['state'], 'district': inputs['district'], 'block': inputs['block']})
        self.app.update_history("mis_state", inputs['state']); self.app.update_history("mis_district", inputs['district']); self.app.update_history("mis_block", inputs['block'])
        
        today_str = datetime.now().strftime("%d-%m-%Y")
        initial_filename = f"MIS_Reports_{today_str}.xlsx"
        
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")], title="Save MIS Reports As", initialdir=self.app.get_user_downloads_path(), initialfile=initial_filename)
        if not save_path: return
        
        self.app.start_automation_thread(self.automation_key, self.run_automation_logic, args=(inputs, save_path))

    def _solve_captcha(self, driver, wait):
        self.app.log_message(self.log_display, "Attempting to solve CAPTCHA...")
        captcha_label_id = "ContentPlaceHolder1_lblStopSpam"; captcha_textbox_id = "ContentPlaceHolder1_txtCaptcha"; verify_button_id = "ContentPlaceHolder1_btnLogin"
        captcha_text = wait.until(EC.presence_of_element_located((By.ID, captcha_label_id))).text
        match = re.search(r'(\d+)\s*([+\-*])\s*(\d+)', captcha_text)
        if not match: raise ValueError("Could not parse CAPTCHA expression.")
        num1, operator, num2 = match.groups(); num1, num2 = int(num1), int(num2)
        result = 0
        if operator == '+': result = num1 + num2
        elif operator == '-': result = num1 - num2
        elif operator == '*': result = num1 * num2
        self.app.log_message(self.log_display, f"Solved: {captcha_text.strip()} = {result}")
        driver.find_element(By.ID, captcha_textbox_id).send_keys(str(result))
        driver.find_element(By.ID, verify_button_id).click()
        return True

    def run_automation_logic(self, inputs, save_path):
        self.app.after(0, self.set_ui_state, True); self.app.clear_log(self.log_display); self.app.log_message(self.log_display, "Starting MIS Report generation...")
        try:
            driver = self.app.get_driver();
            if not driver: return
            wait = WebDriverWait(driver, 20)
            
            total_reports = len(inputs['reports'])
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                for i, report_name in enumerate(inputs['reports']):
                    if self.app.stop_events[self.automation_key].is_set(): self.app.log_message(self.log_display, "Stop signal received.", "warning"); break
                    
                    self.app.after(0, self.update_status, f"Processing {report_name}", (i+1)/total_reports)
                    self.app.log_message(self.log_display, f"--- Processing report {i+1}/{total_reports}: {report_name} ---")
                    
                    try:
                        self.app.log_message(self.log_display, "Navigating to portal and solving CAPTCHA...")
                        driver.get("https://nreganarep.nic.in/netnrega/MISreport4.aspx")
                        self._solve_captcha(driver, wait)
                        self.app.log_message(self.log_display, "CAPTCHA verified. Selecting state...")
                        state_dropdown = wait.until(EC.element_to_be_clickable((By.ID, "ContentPlaceHolder1_ddl_States")))
                        Select(state_dropdown).select_by_visible_text(inputs['state'].upper())
                        wait.until(EC.presence_of_element_located((By.LINK_TEXT, "Dashboard for Delay Monitoring System")))
                        self.app.log_message(self.log_display, f"Finding and scrolling to '{report_name}'...")
                        report_link = wait.until(EC.presence_of_element_located((By.LINK_TEXT, report_name.strip())))
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", report_link)
                        time.sleep(1); report_link.click()
                        if "Aadhaar Status" in report_name:
                            self.app.log_message(self.log_display, "Handling special case, selecting State again...")
                            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['state'].upper()))).click()
                        self.app.log_message(self.log_display, f"Drilling down to District: {inputs['district']}")
                        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['district'].upper()))).click()
                        if "Rejected Wage" in report_name:
                            self.app.log_message(self.log_display, "Handling special case for Rejected Wage report...")
                            block_row = wait.until(EC.presence_of_element_located((By.XPATH, f"//td[normalize-space()='{inputs['block'].upper()}']/ancestor::tr")))
                            block_row.find_element(By.XPATH, ".//td[5]/a").click()
                        else:
                            self.app.log_message(self.log_display, f"Drilling down to Block: {inputs['block']}")
                            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, inputs['block'].upper()))).click()
                        
                        self.app.log_message(self.log_display, "Final page reached. Reading table...")
                        time.sleep(2)
                        
                        try:
                            df_list = pd.read_html(driver.page_source, header=[0, 1])
                            report_df = df_list[-1]
                            report_df.columns = [col[1] for col in report_df.columns]
                            if not report_df.empty and str(report_df.iloc[0, 0]).strip() == '1' and str(report_df.iloc[0, 1]).strip().startswith('2'):
                                self.app.log_message(self.log_display, "Detected and removed junk numeric header row from data.", "warning")
                                report_df = report_df.iloc[1:].reset_index(drop=True)
                        except ValueError:
                            self.app.log_message(self.log_display, "Could not parse multi-level header. Trying single header.", "warning")
                            df_list = pd.read_html(driver.page_source, header=0)
                            report_df = df_list[-1]

                        sheet_name = re.sub(r'[\\/*?:\[\]]', '', report_name)[:30]
                        report_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                        worksheet = writer.sheets[sheet_name]

                        # --- START: FINAL EXCEL FORMATTING ---
                        # 1. Page Setup for Printing (A4, Dynamic Orientation)
                        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE if worksheet.max_column > 7 else worksheet.ORIENTATION_PORTRAIT
                        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
                        worksheet.page_setup.fitToWidth = 1
                        worksheet.page_setup.fitToHeight = 0 
                        worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

                        # 2. Define Styles
                        title_font = Font(bold=True, size=14)
                        header_font = Font(bold=True)
                        total_font = Font(bold=True)
                        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

                        # 3. Format Title (Row 1)
                        worksheet['A1'] = report_name
                        worksheet['A1'].font = title_font
                        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                        if not report_df.empty:
                            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)

                        # 4. Format Header, Data, and Total Rows
                        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                            is_total_row = row[0].value and 'total' in str(row[0].value).lower()
                            for cell in row:
                                if row_idx == 2: # Header row
                                    cell.font = header_font
                                elif is_total_row: # Total row
                                    cell.font = total_font
                                cell.alignment = center_align

                        # 5. Auto-adjust column widths
                        for col_idx in range(1, worksheet.max_column + 1):
                            column_letter = get_column_letter(col_idx)
                            max_length = 0
                            for cell in worksheet[column_letter]:
                                if cell.row == 1: continue
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except: pass
                            adjusted_width = min((max_length + 2), 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        # --- END: FINAL EXCEL FORMATTING ---

                        self.app.log_message(self.log_display, f"Successfully saved and formatted '{report_name}' to sheet '{sheet_name}'.", "success")
                        details = f"Saved to sheet: {sheet_name}"
                        self.app.after(0, lambda r=report_name, d=details: self.results_tree.insert("", "end", values=(r, "Success", d)))

                    except Exception as e:
                        error_msg = str(e).split('\n')[0]
                        self.app.log_message(self.log_display, f"Failed to process '{report_name}': {error_msg}", "error")
                        self.app.after(0, lambda r=report_name, d=error_msg: self.results_tree.insert("", "end", values=(r, "Failed", d), tags=('failed',)))
            
            self.app.log_message(self.log_display, f"Process complete. Excel file saved at: {save_path}", "success")
            
        except Exception as e:
            error_msg = str(e).split('\n')[0]; self.app.log_message(self.log_display, f"A critical error occurred: {error_msg}", "error"); messagebox.showerror("Critical Error", error_msg)
        finally:
            self.app.after(0, self.set_ui_state, False); self.app.after(0, self.update_status, "Automation Finished", 1.0); self.app.after(0, self.app.set_status, "Automation Finished")
            if not self.app.stop_events[self.automation_key].is_set():
                self.app.after(100, lambda: messagebox.showinfo("Complete", f"MIS Report generation has finished.\nFile saved to: {save_path}"))

    def save_inputs(self, inputs):
        try:
            with open(self.config_file, 'w') as f:
                json.dump(inputs, f, indent=4)
        except Exception as e:
            print(f"Error saving MIS inputs: {e}")

    def load_inputs(self):
        if not os.path.exists(self.config_file): return
        try:
            with open(self.config_file, 'r') as f:
                data = json.load(f)
            self.state_entry.delete(0, tkinter.END)
            self.state_entry.insert(0, data.get('state', ''))
            self.district_entry.delete(0, tkinter.END)
            self.district_entry.insert(0, data.get('district', ''))
            self.block_entry.delete(0, tkinter.END)
            self.block_entry.insert(0, data.get('block', ''))
        except Exception as e:
            print(f"Error loading MIS inputs: {e}")